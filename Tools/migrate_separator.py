# -*- coding: utf-8 -*-
"""
migrate_separator.py - Replace struct field separator from '-' to '~' in Excel configs
One-time migration tool. Compatible with Python 2.7+
Usage: python migrate_separator.py [--project-root PATH] [--dry-run]
"""
from __future__ import print_function, unicode_literals
import os
import re
import sys
import argparse

from openpyxl import load_workbook


def to_text(val):
    if val is None:
        return u''
    if isinstance(val, bytes):
        return val.decode('utf-8')
    return type(u'')(val)


def is_struct_type(type_str):
    """Check if a type string is a struct array type (inline or reference)."""
    ts = to_text(type_str).strip()
    if not ts:
        return False
    # Inline struct: StructName{fields}[]
    if re.match(r'^\w+\{.+\}\[\]$', ts):
        return True
    # Reference struct: StructName[] (but not primitive arrays)
    if re.match(r'^(\w+)\[\]$', ts):
        name = re.match(r'^(\w+)\[\]$', ts).group(1)
        if name not in ('int', 'float', 'string', 'bool'):
            return True
    return False


def migrate_item(item):
    """Migrate a single struct item (one element within |-separated cell).

    Rules:
      - A '-' preceded by a word char (including CJK), digit, comma, period, or ')'
        is a field separator -> replace with '~'
      - A '-' at start of item followed by non-'-' char is a separator
        for empty first field -> replace with '~'
      - Other '-' (e.g. negative sign) are kept.

    Examples:
      '131-5000'    -> '131~5000'
      '131--5000'   -> '131~-5000'    (negative value)
      '-0-0-text'   -> '~0~0~text'    (empty first field)
    """
    if not item:
        return item
    # Step 1: replace '-' preceded by a word/digit/comma/period/paren character
    result = re.sub(r'(?<=[\w.,)])-', u'~', item, flags=re.UNICODE)
    # Step 2: handle leading '-' as empty-first-field separator
    if result.startswith('-') and len(result) > 1 and result[1] != '-':
        result = '~' + result[1:]
    return result


def migrate_cell(cell_value):
    """Replace '-' separators with '~' in a struct cell value.
    Process each |-separated item independently.
    """
    text = to_text(cell_value).strip()
    if not text:
        return text

    items = text.split('|')
    new_items = []
    for item in items:
        new_items.append(migrate_item(item))
    return '|'.join(new_items)


def main():
    parser = argparse.ArgumentParser(description='Migrate struct separator from - to ~')
    parser.add_argument('--project-root', default=None)
    parser.add_argument('--dry-run', action='store_true', help='Preview changes without saving')
    args = parser.parse_args()

    if args.project_root:
        project_root = args.project_root
    else:
        project_root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))

    configs_dir = os.path.join(project_root, 'Configs')
    xlsx_files = sorted([f for f in os.listdir(configs_dir) if f.endswith('.xlsx') and not f.startswith('~')])

    if not xlsx_files:
        print("No .xlsx files found in %s" % configs_dir)
        return

    total_changes = 0

    for xlsx_file in xlsx_files:
        fp = os.path.join(configs_dir, xlsx_file)
        wb = load_workbook(fp)
        file_changes = 0

        for ws_name in wb.sheetnames:
            ws = wb[ws_name]
            all_rows = list(ws.iter_rows(min_row=1, values_only=False))
            if len(all_rows) < 3:
                continue

            # Row 2 is types
            types_row = all_rows[1]
            struct_cols = []
            for ci, cell in enumerate(types_row):
                if cell.value and is_struct_type(cell.value):
                    struct_cols.append(ci)

            if not struct_cols:
                continue

            # Process data rows (row 3+)
            for ri in range(2, len(all_rows)):
                row = all_rows[ri]
                for ci in struct_cols:
                    if ci >= len(row):
                        continue
                    cell = row[ci]
                    if cell.value is None:
                        continue
                    old_val = to_text(cell.value).strip()
                    if not old_val or '-' not in old_val:
                        continue
                    new_val = migrate_cell(old_val)
                    if new_val != old_val:
                        if args.dry_run:
                            msg = u"  [%s] %s R%dC%d: '%s' -> '%s'" % (
                                xlsx_file, ws_name, ri + 1, ci + 1, old_val, new_val)
                            if sys.version_info[0] < 3:
                                msg = msg.encode('utf-8')
                            print(msg)
                        cell.value = new_val
                        file_changes += 1

        if file_changes > 0:
            if not args.dry_run:
                wb.save(fp)
            print("%s: %d cells updated" % (xlsx_file, file_changes))
            total_changes += file_changes
        else:
            print("%s: no struct separator found" % xlsx_file)

        wb.close()

    print("\nTotal: %d cells %s" % (total_changes, "would be updated (dry-run)" if args.dry_run else "updated"))


if __name__ == '__main__':
    main()
