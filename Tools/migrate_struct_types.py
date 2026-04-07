# -*- coding: utf-8 -*-
"""
migrate_struct_types.py
One-time migration: remove named struct type names from Excel type declarations.
e.g. BuffSpecialEvent{int type;int[] args}[] -> {int type;int[] args}[]
AttrEntry columns are left unchanged.
"""
import os
import re
import sys
from openpyxl import load_workbook

try:
    string_types = (str, unicode, bytes)
except NameError:
    string_types = (str, bytes)

EXCEL_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), '..', 'Configs')

# Pattern: NamedStruct{fields}[]  ->  capture name and fields
NAMED_STRUCT_RE = re.compile(r'^(\w+)\{(.+)\}\[\]$')

# Types to keep as named (will NOT be converted to anonymous)
KEEP_NAMED = {'AttrEntry'}


def migrate_excel(filepath):
    wb = load_workbook(filepath)
    changed = False
    fname = os.path.basename(filepath)

    for ws_name in wb.sheetnames:
        ws = wb[ws_name]
        # Row 2 is the types row (1-indexed in openpyxl)
        types_row_idx = 2
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=types_row_idx, column=col_idx)
            val = cell.value
            if not val or not isinstance(val, string_types):
                continue
            if isinstance(val, bytes):
                val = val.decode('utf-8')
            val = val.strip()

            m = NAMED_STRUCT_RE.match(val)
            if m:
                struct_name = m.group(1)
                fields_str = m.group(2)
                if struct_name in KEEP_NAMED:
                    continue
                new_val = '{%s}[]' % fields_str
                cell.value = new_val
                field_name = ''
                name_cell = ws.cell(row=1, column=col_idx)
                if name_cell.value:
                    field_name = str(name_cell.value).strip()
                msg = '  [%s] %s.%s: %s -> %s' % (fname, ws_name, field_name, val, new_val)
                if sys.version_info[0] < 3:
                    print(msg.encode('utf-8'))
                else:
                    print(msg)
                changed = True

    if changed:
        wb.save(filepath)
        print('  Saved %s' % fname)
    else:
        print('  No changes in %s' % fname)
    wb.close()


def main():
    print('Migrating struct type declarations in Excel files ...')
    print('Excel dir: %s' % EXCEL_DIR)
    print('')

    xlsx_files = sorted([f for f in os.listdir(EXCEL_DIR) if f.endswith('.xlsx') and not f.startswith('~')])
    total_changes = 0

    for xlsx_file in xlsx_files:
        fp = os.path.join(EXCEL_DIR, xlsx_file)
        migrate_excel(fp)

    print('')
    print('Done.')


if __name__ == '__main__':
    main()
