# -*- coding: utf-8 -*-
"""
config_gen.py - Excel -> JSON + C# config generator for GeometryTD
Usage: python config_gen.py [--project-root PATH]
Compatible with Python 2.7+
"""
from __future__ import print_function, unicode_literals
import os
import re
import json
import sys
import codecs
import argparse
from collections import OrderedDict

from openpyxl import load_workbook

# Set default encoding to utf-8 for Python 2
if sys.version_info[0] < 3:
    reload(sys)
    sys.setdefaultencoding('utf-8')

# ---------------------------------------------------------------------------
# Type System
# ---------------------------------------------------------------------------

class TypeInfo(object):
    pass

class PrimitiveType(TypeInfo):
    def __init__(self, name):
        self.name = name  # 'int', 'float', 'string', 'bool'

    @property
    def csharp_type(self):
        return self.name

class ArrayType(TypeInfo):
    def __init__(self, element_type):
        self.element_type = element_type  # PrimitiveType

    @property
    def csharp_type(self):
        return "%s[]" % self.element_type.csharp_type

class StructArrayType(TypeInfo):
    def __init__(self, struct_name, fields, is_anonymous=False):
        self.struct_name = struct_name
        self.fields = fields  # list of (TypeInfo, field_name) or None for ref
        self.is_anonymous = is_anonymous

    @property
    def csharp_type(self):
        return "%s[]" % self.struct_name


def parse_struct_fields(fields_str):
    """Parse 'int id;int value' -> [(PrimitiveType('int'), 'id'), ...]"""
    fields = []
    for field_def in fields_str.split(';'):
        field_def = field_def.strip()
        if not field_def:
            continue
        parts = field_def.rsplit(' ', 1)
        if len(parts) != 2:
            raise ValueError("Invalid field definition: %s" % field_def)
        type_str, field_name = parts[0].strip(), parts[1].strip()
        field_type = parse_type(type_str)
        fields.append((field_type, field_name))
    return fields


def parse_type(type_str):
    """Parse type declaration string -> TypeInfo"""
    type_str = type_str.strip()

    # Anonymous struct: {fields}[]
    m = re.match(r'^\{(.+)\}\[\]$', type_str)
    if m:
        return StructArrayType(None, parse_struct_fields(m.group(1)), is_anonymous=True)

    # StructName{fields}[]
    m = re.match(r'^(\w+)\{(.+)\}\[\]$', type_str)
    if m:
        return StructArrayType(m.group(1), parse_struct_fields(m.group(2)))

    # TypeName[]
    m = re.match(r'^(\w+)\[\]$', type_str)
    if m:
        name = m.group(1)
        if name in ('int', 'float', 'string', 'bool'):
            return ArrayType(PrimitiveType(name))
        return StructArrayType(name, None)  # reference

    if type_str in ('int', 'float', 'string', 'bool'):
        return PrimitiveType(type_str)

    raise ValueError("Unknown type: %s" % type_str)


# ---------------------------------------------------------------------------
# Value Conversion
# ---------------------------------------------------------------------------

def get_default_value(type_info):
    if isinstance(type_info, PrimitiveType):
        return {'int': 0, 'float': 0.0, 'string': '', 'bool': False}[type_info.name]
    if isinstance(type_info, (ArrayType, StructArrayType)):
        return []
    return None


def to_text(val):
    """Safely convert any value to unicode text (Python 2/3)."""
    if val is None:
        return u''
    if isinstance(val, bytes):
        return val.decode('utf-8')
    return type(u'')(val)


def convert_primitive(raw, ptype_name):
    if ptype_name == 'int':
        return int(float(raw)) if raw != '' else 0
    if ptype_name == 'float':
        return float(raw) if raw != '' else 0.0
    if ptype_name == 'string':
        return to_text(raw)
    if ptype_name == 'bool':
        if isinstance(raw, bool):
            return raw
        return to_text(raw).lower() in ('true', '1')
    return raw


def convert_value(raw, type_info, shared_types):
    """Convert Excel cell value to Python value for JSON."""
    if raw is None:
        return get_default_value(type_info)
    raw_str = to_text(raw).strip() if not isinstance(raw, bool) else raw

    if isinstance(raw, bool) and isinstance(type_info, PrimitiveType) and type_info.name == 'bool':
        return raw

    if isinstance(type_info, PrimitiveType):
        if raw_str == '' or raw_str == u'':
            return get_default_value(type_info)
        return convert_primitive(raw_str, type_info.name)

    if isinstance(type_info, ArrayType):
        if raw_str == '' or raw_str == u'' or raw_str == '0' or raw_str == u'0':
            return []
        parts = [p.strip() for p in to_text(raw_str).split('|') if p.strip()]
        return [convert_primitive(p, type_info.element_type.name) for p in parts]

    if isinstance(type_info, StructArrayType):
        if raw_str == '' or raw_str == u'' or raw_str == '0' or raw_str == u'0':
            return []
        fields = type_info.fields
        if fields is None:
            fields = shared_types.get(type_info.struct_name)
            if fields is None:
                raise ValueError("Unknown struct type: %s" % type_info.struct_name)
        items = []
        for item_str in to_text(raw_str).split('|'):
            item_str = item_str.strip()
            if not item_str:
                continue
            field_values = item_str.split('~')
            obj = OrderedDict()
            for i, (ft, fn) in enumerate(fields):
                if i < len(field_values):
                    fv = field_values[i].strip()
                    if isinstance(ft, PrimitiveType):
                        obj[fn] = convert_primitive(fv, ft.name)
                    elif isinstance(ft, ArrayType):
                        if fv:
                            inner = [p.strip() for p in fv.split('#') if p.strip()]
                            obj[fn] = [convert_primitive(p, ft.element_type.name) for p in inner]
                        else:
                            obj[fn] = []
                    else:
                        obj[fn] = fv
                else:
                    obj[fn] = get_default_value(ft)
            items.append(obj)
        return items

    return raw


# ---------------------------------------------------------------------------
# Excel Parsing
# ---------------------------------------------------------------------------

def parse_sheet(ws, shared_types):
    """Parse a worksheet -> (fields_info, rows_data)
    fields_info: list of (field_name, TypeInfo)
    rows_data: list of OrderedDict
    """
    all_rows = list(ws.iter_rows(min_row=1, values_only=True))
    if len(all_rows) < 2:
        return [], []

    names_row = all_rows[0]
    types_row = all_rows[1]
    data_rows = all_rows[2:]

    fields_info = []
    valid_cols = []
    for ci in range(len(names_row)):
        fname = names_row[ci]
        ftype_str = types_row[ci] if ci < len(types_row) else None
        if not fname or not ftype_str:
            continue
        fname = str(fname).strip()
        ftype_str = str(ftype_str).strip()
        if not fname or not ftype_str:
            continue
        ti = parse_type(ftype_str)
        # Anonymous struct: derive inner class name from field name
        if isinstance(ti, StructArrayType) and ti.is_anonymous:
            ti.struct_name = fname[0].upper() + fname[1:] + 'Item'
        # Named struct: register to shared_types for reference resolution
        elif isinstance(ti, StructArrayType) and ti.fields is not None:
            shared_types[ti.struct_name] = ti.fields
        fields_info.append((fname, ti))
        valid_cols.append(ci)

    rows_data = []
    for dr in data_rows:
        # Skip empty rows
        if not dr or all(c is None for c in dr):
            continue
        row_obj = OrderedDict()
        for idx, ci in enumerate(valid_cols):
            fname, ti = fields_info[idx]
            raw = dr[ci] if ci < len(dr) else None
            row_obj[fname] = convert_value(raw, ti, shared_types)
        rows_data.append(row_obj)

    return fields_info, rows_data


def parse_excel(file_path, shared_types):
    """Parse an xlsx file.
    Returns: list of (base_name, list_fields, list_data, meta_fields, meta_data) - one per sheet
    """
    wb = load_workbook(file_path, read_only=True, data_only=True)
    file_base_name = os.path.splitext(os.path.basename(file_path))[0]  # e.g. hero_config
    file_short_name = file_base_name.replace('_config', '')  # e.g. hero

    results = []

    for ws_name in wb.sheetnames:
        ws = wb[ws_name]
        # Remove Chinese characters from sheet name for processing
        sn_clean = remove_chinese(ws_name).strip().lower()
        sn = sn_clean
        
        # Skip meta sheets
        if sn.endswith('_meta') or sn == 'meta':
            continue
            
        # Determine base name from sheet name
        # If sheet name looks like a config name (contains '_config' or is descriptive), use it
        # Otherwise, use the file name
        if sn and sn != 'list' and sn != file_short_name and sn != file_base_name and sn != file_base_name.replace('_', ''):
            # Sheet name is descriptive, use it as config name
            if sn.endswith('_config'):
                base_name = sn
            elif '_config' in sn:
                # Handle cases like 'hero_config_sheet1' -> 'hero_config'
                base_name = sn.split('_config')[0] + '_config'
            else:
                base_name = sn + '_config'
        else:
            # Sheet name is generic (like 'list' or matches file name), use file name
            base_name = file_base_name
            
        list_fields, list_data = parse_sheet(ws, shared_types)
        
        # Only add if there's actual data or fields
        if list_fields or list_data:
            results.append((base_name, list_fields, list_data, [], None))

    wb.close()
    return results


# ---------------------------------------------------------------------------
# Name Helpers
# ---------------------------------------------------------------------------

def remove_chinese(text):
    """Remove Chinese characters from text."""
    # Chinese characters are in the range \u4e00-\u9fff
    return re.sub(u'[\u4e00-\u9fff]', '', text)

def to_pascal(snake):
    """hero_config -> HeroConfig, bullet_style -> BulletStyle"""
    return ''.join(w.capitalize() for w in snake.split('_'))


def to_camel(snake):
    """hero -> hero, bullet_event -> bulletEvent"""
    parts = snake.split('_')
    return parts[0] + ''.join(w.capitalize() for w in parts[1:])


def get_class_names(base_name):
    """Given 'hero_config', return (HeroConfig, HeroMeta, HeroConfigData, hero, Hero, heroTable)"""
    short = base_name.replace('_config', '')
    pascal = to_pascal(short)
    camel = to_camel(short)
    return (
        pascal + 'Config',     # HeroConfig
        pascal + 'Meta',       # HeroMeta
        pascal + 'ConfigData', # HeroConfigData
        short,                 # hero (snake_case)
        pascal,                # Hero (PascalCase)
        camel + 'Table',       # heroTable (camelCase)
    )


# ---------------------------------------------------------------------------
# JSON Generation
# ---------------------------------------------------------------------------

def generate_json(base_name, list_data, output_dir):
    json_obj = OrderedDict()
    if list_data is not None and len(list_data) > 0:
        json_obj['items'] = list_data

    # Handle Unicode base_name for file path
    if sys.version_info[0] < 3 and isinstance(base_name, unicode):
        base_name = base_name.encode('utf-8')
    
    path = os.path.join(output_dir, base_name + '.json')
    with codecs.open(path, 'w', encoding='utf-8') as f:
        json.dump(json_obj, f, ensure_ascii=False, indent=2, sort_keys=False)
    print("  JSON -> %s" % path)


# ---------------------------------------------------------------------------
# C# Generation
# ---------------------------------------------------------------------------

HEADER = """// ==========================================================
// AUTO-GENERATED by config_gen.py - DO NOT EDIT MANUALLY
// ==========================================================
"""


def csharp_type_str(ti):
    if isinstance(ti, PrimitiveType):
        return ti.name
    if isinstance(ti, ArrayType):
        return "%s[]" % ti.element_type.csharp_type
    if isinstance(ti, StructArrayType):
        return "%s[]" % ti.struct_name
    return "object"


def generate_config_cs(base_name, list_fields, output_dir):
    config_cls, meta_cls, data_cls, short, pascal, table_field = get_class_names(base_name)
    lines = [HEADER]
    lines.append("using System;")
    lines.append("using System.Collections.Generic;")
    lines.append("")
    lines.append("namespace GeometryTD")
    lines.append("{")

    has_list = len(list_fields) > 0

    # Config class (list item)
    if has_list:
        # Collect anonymous struct types that need inner class generation
        list_inner = []
        for fname, ti in list_fields:
            if isinstance(ti, StructArrayType) and ti.is_anonymous:
                list_inner.append((ti.struct_name, ti.fields))

        lines.append("    [Serializable]")
        lines.append("    public class %s" % config_cls)
        lines.append("    {")
        # Generate inner classes first
        for ic_name, ic_fields in list_inner:
            lines.append("        [Serializable]")
            lines.append("        public class %s" % ic_name)
            lines.append("        {")
            for ft, fn in ic_fields:
                lines.append("            public %s %s;" % (csharp_type_str(ft), fn))
            lines.append("        }")
            lines.append("")
        # Then generate fields
        for fname, ti in list_fields:
            lines.append("        public %s %s;" % (csharp_type_str(ti), fname))
        lines.append("    }")
        lines.append("")

    # Data wrapper
    lines.append("    [Serializable]")
    lines.append("    public class %s" % data_cls)
    lines.append("    {")
    if has_list:
        lines.append("        public List<%s> items;" % config_cls)
    lines.append("    }")

    lines.append("}")
    lines.append("")

    # Handle Unicode base_name for file path
    if sys.version_info[0] < 3 and isinstance(base_name, unicode):
        base_name = base_name.encode('utf-8')
    
    path = os.path.join(output_dir, config_cls + '.cs')
    with codecs.open(path, 'w', encoding='utf-8') as f:
        f.write('\n'.join(lines))
    print("  C# -> %s" % path)



def generate_cfg_cs(tables, output_dir):
    """Generate Cfg.cs - static shortcut entry point."""
    lines = [HEADER]
    lines.append("namespace GeometryTD")
    lines.append("{")
    lines.append("    public static class Cfg")
    lines.append("    {")
    lines.append("        private static ConfigManager M { get { return ConfigManager.Instance; } }")
    lines.append("")

    for t in tables:
        config_cls, meta_cls, data_cls, short, pascal, table_field = get_class_names(t['base_name'])
        has_list = t['has_list']

        if has_list:
            lines.append("        public static ConfigTable<%s> %s { get { return M.%s; } }"
                          % (config_cls, pascal, table_field))

    lines.append("    }")
    lines.append("}")
    lines.append("")

    path = os.path.join(output_dir, 'Cfg.cs')
    with codecs.open(path, 'w', encoding='utf-8') as f:
        f.write('\n'.join(lines))
    print("  C# -> %s" % path)


# ---------------------------------------------------------------------------
# ConfigManager Generation
# ---------------------------------------------------------------------------

def extract_user_code(file_path):
    """Extract USER CODE sections from existing ConfigManager.cs"""
    sections = {}
    if not os.path.exists(file_path):
        return sections
    with codecs.open(file_path, 'r', encoding='utf-8') as f:
        content = f.read()

    pattern = r'// USER CODE START - (\w+)\s*\n(.*?)\s*// USER CODE END - \1'
    for m in re.finditer(pattern, content, re.DOTALL):
        sections[m.group(1)] = m.group(2)#.strip()
    return sections


def generate_config_manager_cs(tables, output_dir, existing_path=None):
    user_code = {}
    if existing_path:
        user_code = extract_user_code(existing_path)

    lines = [HEADER]
    lines.append("using System;")
    lines.append("using System.Collections.Generic;")
    lines.append("using UnityEngine;")
    lines.append("")
    lines.append("namespace GeometryTD")
    lines.append("{")
    lines.append("    public class ConfigManager : MonoBehaviour")
    lines.append("    {")
    lines.append("        public static ConfigManager Instance { get; private set; }")
    lines.append("")

    # Table fields
    lines.append("        // --- AUTO-GENERATED TABLE FIELDS ---")
    for t in tables:
        config_cls, meta_cls, data_cls, short, pascal, table_field = get_class_names(t['base_name'])
        has_list = t['has_list']
        key_field = t['key_field']
        json_name = t['base_name']

        if has_list:
            lines.append("        public ConfigTable<%s> %s;" % (config_cls, table_field))
    lines.append("")

    # USER CODE - Fields
    lines.append("        // USER CODE START - Fields")
    lines.append(user_code.get('Fields', ''))
    lines.append("        // USER CODE END - Fields")
    lines.append("")

    # Awake
    lines.append("        private void Awake()")
    lines.append("        {")
    lines.append("            if (Instance != null && Instance != this) { Destroy(gameObject); return; }")
    lines.append("            Instance = this;")
    lines.append("            DontDestroyOnLoad(gameObject);")
    lines.append("            LoadAllConfigs();")
    lines.append("        }")
    lines.append("")

    # LoadAllConfigs
    lines.append("        private void LoadAllConfigs()")
    lines.append("        {")
    for t in tables:
        config_cls, meta_cls, data_cls, short, pascal, table_field = get_class_names(t['base_name'])
        has_list = t['has_list']
        key_field = t['key_field']
        json_name = t['base_name']

        lines.append("            {")
        lines.append('                var data = LoadConfig<%s>("Configs/%s");' % (data_cls, json_name))
        if has_list:
            lines.append("                %s = new ConfigTable<%s>();" % (table_field, config_cls))
            lines.append("                %s.Init(data.items, c => c.%s);" % (table_field, key_field))
        lines.append("            }")

    lines.append("")
    lines.append("            // USER CODE START - AfterLoad")
    lines.append(user_code.get('AfterLoad', ''))
    lines.append("            // USER CODE END - AfterLoad")
    lines.append("        }")
    lines.append("")

    # LoadConfig helper
    lines.append("        private T LoadConfig<T>(string path)")
    lines.append("        {")
    lines.append("            TextAsset textAsset = Resources.Load<TextAsset>(path);")
    lines.append("            if (textAsset == null)")
    lines.append("            {")
    lines.append('                Debug.LogError("[ConfigManager] Failed to load: " + path);')
    lines.append("                return default(T);")
    lines.append("            }")
    lines.append("            T config = JsonUtility.FromJson<T>(textAsset.text);")
    lines.append("            if (config == null)")
    lines.append("            {")
    lines.append('                Debug.LogError("[ConfigManager] Failed to parse: " + path);')
    lines.append("                return default(T);")
    lines.append("            }")
    lines.append("            return config;")
    lines.append("        }")
    lines.append("")

    # USER CODE - CustomMethods
    lines.append("        // USER CODE START - CustomMethods")
    lines.append(user_code.get('CustomMethods', ''))
    lines.append("        // USER CODE END - CustomMethods")

    lines.append("    }")
    lines.append("}")
    lines.append("")

    path = os.path.join(output_dir, 'ConfigManager.cs')
    with codecs.open(path, 'w', encoding='utf-8') as f:
        f.write('\n'.join(lines))
    print("  C# -> %s" % path)


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(description='Generate JSON + C# from Excel configs')
    parser.add_argument('--project-root', default=None,
                        help='Project root directory (default: parent of Tools/)')
    args = parser.parse_args()

    if args.project_root:
        project_root = args.project_root
    else:
        project_root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))

    configs_dir = os.path.join(project_root, 'Configs')
    json_dir = os.path.join(project_root, 'Assets', 'Resources', 'Configs')
    cs_config_dir = os.path.join(project_root, 'Assets', 'Scripts', 'Data', 'Configs')
    cs_data_dir = os.path.join(project_root, 'Assets', 'Scripts', 'Data')
    cs_core_dir = os.path.join(project_root, 'Assets', 'Scripts', 'Core')
    existing_cm = os.path.join(cs_core_dir, 'ConfigManager.cs')

    for d in [json_dir, cs_config_dir, cs_data_dir, cs_core_dir]:
        if not os.path.exists(d):
            os.makedirs(d)

    xlsx_files = sorted([f for f in os.listdir(configs_dir) if f.endswith('.xlsx') and not f.startswith('~')])
    if not xlsx_files:
        print("No .xlsx files found in %s" % configs_dir)
        return

    print("Found %d Excel files:" % len(xlsx_files))
    for f in xlsx_files:
        print("  %s" % f)
    print("")

    shared_types = OrderedDict()
    tables = []

    # Pre-scan: collect all struct type definitions from type rows
    print("Pre-scanning struct types ...")
    for xlsx_file in xlsx_files:
        fp = os.path.join(configs_dir, xlsx_file)
        wb = load_workbook(fp, read_only=True, data_only=True)
        for ws_name in wb.sheetnames:
            ws = wb[ws_name]
            all_rows = list(ws.iter_rows(min_row=1, max_row=2, values_only=True))
            if len(all_rows) >= 2:
                types_row = all_rows[1]
                for cell in types_row:
                    if cell:
                        ts = to_text(cell).strip()
                        m = re.match(r'^(\w+)\{(.+)\}\[\]$', ts)
                        if m:
                            sname = m.group(1)
                            if sname not in shared_types:
                                shared_types[sname] = parse_struct_fields(m.group(2))
        wb.close()
    print("  Found %d shared struct types: %s" % (len(shared_types), ', '.join(shared_types.keys())))
    print("")

    # Pass 1: parse all excels with complete shared_types
    for xlsx_file in xlsx_files:
        fp = os.path.join(configs_dir, xlsx_file)
        print("Parsing %s ..." % xlsx_file)
        excel_results = parse_excel(fp, shared_types)
        
        for base_name, list_fields, list_data, meta_fields, meta_data in excel_results:
            # Determine key field (first field of list, or empty)
            key_field = ''
            if list_fields:
                key_field = list_fields[0][0]

            tables.append({
                'base_name': base_name,
                'list_fields': list_fields,
                'list_data': list_data,
                'has_list': len(list_data) > 0 or len(list_fields) > 0,
                'key_field': key_field,
            })

    print("")

    # Pass 2: generate outputs
    for t in tables:
        base_name = t['base_name']
        print("Generating %s ..." % base_name)
        generate_json(base_name, t['list_data'], json_dir)
        generate_config_cs(base_name, t['list_fields'], cs_config_dir)

    print("")
    print("Generating Cfg.cs ...")
    generate_cfg_cs(tables, cs_core_dir)

    print("Generating ConfigManager.cs ...")
    generate_config_manager_cs(tables, cs_core_dir, existing_cm)

    print("")
    print("Done! Generated %d tables." % len(tables))


if __name__ == '__main__':
    main()
