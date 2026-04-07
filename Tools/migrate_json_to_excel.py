# -*- coding: utf-8 -*-
"""
migrate_json_to_excel.py - Convert existing JSON configs to Excel format
One-time migration tool. Compatible with Python 2.7+
Usage: python migrate_json_to_excel.py [--project-root PATH]
"""
from __future__ import print_function, unicode_literals
import os
import json
import sys
import codecs
import argparse
from collections import OrderedDict

from openpyxl import Workbook

# ---------------------------------------------------------------------------
# Schema Definitions: define the type row for each config table
# Format: (field_name, type_declaration_string)
# ---------------------------------------------------------------------------

SCHEMAS = OrderedDict([
    ('hero_config', {
        'list': [
            ('id', 'int'),
            ('name', 'string'),
            ('description', 'string'),
            ('role', 'int'),
            ('attack_skill_ids', 'int[]'),
            ('skill_xp_interval', 'float'),
            ('skill_xp_min', 'int'),
            ('skill_xp_max', 'int'),
            ('attrs', 'AttrEntry{int id;int value}[]'),
            ('charge_buff_ids', 'int[]'),
        ],
        'meta': [
            ('default_hero_id', 'int'),
        ],
        'json_root': 'heroes',
        'meta_from_game_config': {'default_hero_id': 'default_hero_id'},
    }),
    ('monster_config', {
        'list': [
            ('id', 'int'),
            ('name', 'string'),
            ('role', 'int'),
            ('level', 'int'),
            ('is_boss', 'bool'),
            ('is_elite', 'bool'),
            ('attack_skill_ids', 'int[]'),
            ('attack_interval', 'float'),
            ('attrs', 'AttrEntry[]'),
        ],
        'meta': [
            ('boss_monster_id', 'int'),
        ],
        'json_root': 'monsters',
        'meta_from_game_config': {'boss_monster_id': 'boss_monster_id'},
    }),
    ('skill_config', {
        'list': [
            ('id', 'int'),
            ('level', 'int'),
            ('name', 'string'),
            ('des', 'string'),
            ('icon', 'string'),
            ('category', 'string'),
            ('dmg', 'int'),
            ('dmgType', 'int'),
            ('bulletSpeed', 'float'),
            ('cd', 'float'),
            ('bulletStyleId', 'int'),
            ('attack_range', 'float'),
            ('events', 'int[]'),
            ('enemyEvents', 'int[]'),
            ('bulletEvents', 'int[]'),
        ],
        'meta': [
            ('slot_ids', 'int[]'),
        ],
        'json_root': 'skills',
        'meta_from_game_config': {'slot_ids': 'skill_slot_ids'},
    }),
    ('skill_pool_config', {
        'list': [
            ('id', 'int'),
            ('name', 'string'),
            ('desList', 'string[]'),
            ('icon', 'string'),
            ('dragHint', 'string'),
        ],
        'json_root': 'skill_pool_config',
    }),
    ('bullet_style_config', {
        'list': [
            ('id', 'int'),
            ('prefabPath', 'string'),
        ],
        'json_file': 'bullet_config',
        'json_root': 'bulletStyles',
    }),
    ('arcane_config', {
        'list': [
            ('id', 'int'),
            ('name', 'string'),
            ('desList', 'string[]'),
            ('icon', 'string'),
            ('dmg', 'int'),
            ('dmgType', 'int'),
            ('radius', 'float'),
            ('tickInterval', 'float'),
            ('cd', 'float'),
            ('runeCost', 'int'),
            ('runeType', 'int'),
            ('events', 'int[]'),
            ('enemyEvents', 'int[]'),
            ('bulletEvents', 'int[]'),
        ],
        'meta': [
            ('slot_ids', 'int[]'),
        ],
        'json_root': 'arcanes',
        'meta_from_game_config': {'slot_ids': 'arcane_slot_ids'},
    }),
    ('event_config', {
        'list': [
            ('id', 'int'),
            ('type', 'int'),
            ('name', 'string'),
            ('des', 'string'),
            ('args', 'int[]'),
        ],
        'json_root': 'events',
    }),
    ('bullet_event_config', {
        'list': [
            ('id', 'int'),
            ('type', 'int'),
            ('name', 'string'),
            ('des', 'string'),
            ('args', 'int[]'),
        ],
        'json_root': 'bulletEvents',
    }),
    ('buff_config', {
        'list': [
            ('id', 'int'),
            ('name', 'string'),
            ('icon', 'string'),
            ('desc', 'string'),
            ('overlap', 'int'),
            ('probability', 'int'),
            ('lastTime', 'int'),
            ('jumpTime', 'int'),
            ('persistJson', 'string'),
            ('position', 'string'),
            ('type', 'int'),
            ('dispel', 'int'),
            ('attribute', 'AttrEntry[]'),
            ('evtDmgRate', '{int type;int rate}[]'),
            ('evtDamage', 'int[]'),
            ('evtWhenEnd', 'int[]'),
            ('specialEvent', '{int type;int[] args}[]'),
        ],
        'json_root': 'buffs',
    }),
    ('passive_config', {
        'list': [
            ('id', 'int'),
            ('name', 'string'),
            ('icon', 'string'),
            ('des', 'string'),
            ('eventTarget', 'int[]'),
            ('eventRemove', 'int[]'),
            ('eventCond', '{int id;int[] args}[]'),
            ('events', 'int[]'),
        ],
        'json_root': 'passives',
    }),
    ('event_effect_config', {
        'list': [
            ('eventType', 'int'),
            ('duration', 'float'),
            ('target', 'string'),
            ('prefabPath', 'string'),
        ],
        'json_root': 'effects',
    }),
    ('level_config', {
        'list': [
            ('id', 'int'),
            ('name', 'string'),
            ('des', 'string'),
            ('bg', 'string'),
            ('conditions', 'int[]'),
            ('hard', 'int'),
            ('spawn_interval', 'float'),
            ('coinNormalKill', 'int'),
            ('coinEliteKill', 'int'),
            ('coinBossKill', 'int'),
            ('coinSelfDestructRate', 'float'),
            ('monsterList', '{int id;int generate}[]'),
            ('superMList', '{int id;int num;int generate}[]'),
            ('bossList', '{int id;int num}[]'),
        ],
        'json_root': 'levels',
    }),
    ('condition_config', {
        'list': [
            ('id', 'int'),
            ('desc', 'string'),
            ('type', 'int'),
            ('p1', 'int'),
            ('p2', 'int'),
        ],
        'json_root': 'conditions',
    }),
    ('role_config', {
        'list': [
            ('id', 'int'),
            ('name', 'string'),
            ('prefabPath', 'string'),
            ('portraitPath', 'string'),
        ],
        'json_root': 'roles',
    }),
    ('attribute_config', {
        'list': [
            ('id', 'int'),
            ('name', 'string'),
            ('des', 'string'),
            ('type', 'int'),
            ('downLimit', 'int'),
            ('upLimit', 'int'),
            ('powerType', 'int'),
        ],
        'json_root': 'attributes',
    }),
    ('story_collection_config', {
        'list': [
            ('id', 'int'),
            ('name', 'string'),
            ('description', 'string'),
            ('icon', 'string'),
            ('startNodeId', 'int'),
            ('endingNodeIds', 'int[]'),
        ],
        'json_root': 'collections',
    }),
    ('story_node_config', {
        'list': [
            ('id', 'int'),
            ('collectionId', 'int'),
            ('name', 'string'),
            ('icon', 'string'),
            ('type', 'int'),
            ('levelId', 'int'),
            ('bossEvents', '{int dialogueId;int choiceGroupId}[]'),
            ('dialogueId', 'int'),
            ('choiceGroupId', 'int'),
            ('shopId', 'int'),
            ('defaultNextNodeId', 'int'),
            ('failNodeId', 'int'),
            ('endingType', 'int'),
            ('endingCg', 'string'),
            ('branchLineCount', 'int'),
            ('nextNodes', '{int nodeId;int[] conditions}[]'),
        ],
        'json_root': 'nodes',
    }),
    ('dialogue_config', {
        'list': [
            ('id', 'int'),
            ('lines', '{string speaker;int roleId;int portraitSide;string text}[]'),
        ],
        'json_root': 'dialogues',
    }),
    ('choice_group_config', {
        'list': [
            ('id', 'int'),
            ('title', 'string'),
            ('options', '{int id;string text;string description;int effectId;bool triggerBattle;int goldReward}[]'),
        ],
        'json_file': 'choice_option_config',
        'json_root': 'choiceGroups',
    }),
    ('passive_effect_config', {
        'list': [
            ('id', 'int'),
            ('name', 'string'),
            ('description', 'string'),
            ('icon', 'string'),
            ('rarity', 'int'),
            ('effectType', 'int'),
            ('targetAttrId', 'int'),
            ('valueType', 'int'),
            ('value', 'float'),
            ('stackable', 'bool'),
            ('maxStack', 'int'),
        ],
        'json_root': 'effects',
    }),
    ('event_shop_config', {
        'list': [
            ('id', 'int'),
            ('name', 'string'),
            ('refreshCount', 'int'),
            ('items', '{int effectId;int price;int weight}[]'),
        ],
        'json_root': 'shops',
    }),
    ('global_config', {
        'meta': [
            ('kill_count_for_boss', 'int'),
            ('monster_spawn_interval', 'float'),
        ],
        'json_file': 'game_config',
        'meta_from_game_config': {
            'kill_count_for_boss': 'kill_count_for_boss',
            'monster_spawn_interval': 'monster_spawn_interval',
        },
    }),
])


# ---------------------------------------------------------------------------
# Value Encoding (JSON value -> Excel cell format)
# ---------------------------------------------------------------------------

def encode_int_array(arr):
    if not arr:
        return ''
    return '|'.join(to_text(int(v)) for v in arr)


def encode_string_array(arr):
    if not arr:
        return ''
    return '|'.join(to_text(v) for v in arr)


def encode_float_array(arr):
    if not arr:
        return ''
    return '|'.join(to_text(v) for v in arr)


def to_text(val):
    """Safely convert any value to unicode text (Python 2/3)."""
    if val is None:
        return ''
    if isinstance(val, bytes):
        return val.decode('utf-8')
    return type(u'')(val)


def encode_struct_field(val, field_type_str):
    """Encode a single field value inside a struct."""
    if field_type_str.endswith('[]'):
        # inner array, comma separated
        base = field_type_str[:-2]
        if not val:
            return ''
        if base == 'int':
            return ','.join(to_text(int(v)) for v in val)
        return ','.join(to_text(v) for v in val)
    if field_type_str == 'int':
        return to_text(int(val)) if val is not None else '0'
    if field_type_str == 'float':
        return to_text(val) if val is not None else '0'
    if field_type_str == 'bool':
        return 'true' if val else 'false'
    if field_type_str == 'string':
        return to_text(val) if val is not None else ''
    return to_text(val) if val is not None else ''


def parse_struct_def(type_str):
    """Extract field definitions from a struct type string.
    E.g. 'AttrEntry{int id;int value}[]' -> [('int', 'id'), ('int', 'value')]
    or '{int id;int value}[]' -> [('int', 'id'), ('int', 'value')]  (anonymous)
    or 'AttrEntry[]' -> None (reference)
    """
    import re
    # Named struct: Name{fields}[]
    m = re.match(r'^\w+\{(.+)\}\[\]$', type_str)
    if m:
        fields = []
        for fd in m.group(1).split(';'):
            fd = fd.strip()
            if not fd:
                continue
            parts = fd.rsplit(' ', 1)
            fields.append((parts[0].strip(), parts[1].strip()))
        return fields
    # Anonymous struct: {fields}[]
    m = re.match(r'^\{(.+)\}\[\]$', type_str)
    if m:
        fields = []
        for fd in m.group(1).split(';'):
            fd = fd.strip()
            if not fd:
                continue
            parts = fd.rsplit(' ', 1)
            fields.append((parts[0].strip(), parts[1].strip()))
        return fields
    return None


def encode_struct_array(arr, type_str, all_struct_defs):
    """Encode a struct array to Excel cell format."""
    if not arr:
        return ''

    # Get field definitions
    fields = parse_struct_def(type_str)
    if fields is None:
        # Reference type - find from all_struct_defs
        import re
        m = re.match(r'^(\w+)\[\]$', type_str)
        if m:
            sname = m.group(1)
            fields = all_struct_defs.get(sname)
    if fields is None:
        return json.dumps(arr, ensure_ascii=False)

    items = []
    for obj in arr:
        parts = []
        for ftype, fname in fields:
            val = obj.get(fname)
            parts.append(encode_struct_field(val, ftype))
        items.append('~'.join(parts))
    return '|'.join(items)


def encode_value(val, type_str, all_struct_defs):
    """Encode a JSON value to Excel cell string based on type."""
    if type_str == 'int':
        return int(val) if val is not None else 0
    if type_str == 'float':
        return float(val) if val is not None else 0.0
    if type_str == 'string':
        return to_text(val) if val is not None and val != '' else ''
    if type_str == 'bool':
        return bool(val) if val is not None else False
    if type_str == 'int[]':
        return encode_int_array(val) if val else ''
    if type_str == 'float[]':
        return encode_float_array(val) if val else ''
    if type_str == 'string[]':
        return encode_string_array(val) if val else ''
    # Struct array
    if type_str.endswith('[]'):
        return encode_struct_array(val, type_str, all_struct_defs)
    return to_text(val) if val is not None else ''


# ---------------------------------------------------------------------------
# Main Migration
# ---------------------------------------------------------------------------

def collect_struct_defs():
    """Collect all struct definitions from schemas."""
    import re
    defs = {}
    for config_name, schema in SCHEMAS.items():
        for section in ('list', 'meta'):
            fields = schema.get(section, [])
            for fname, ftype in fields:
                m = re.match(r'^(\w+)\{(.+)\}\[\]$', ftype)
                if m:
                    sname = m.group(1)
                    struct_fields = []
                    for fd in m.group(2).split(';'):
                        fd = fd.strip()
                        if not fd:
                            continue
                        parts = fd.rsplit(' ', 1)
                        struct_fields.append((parts[0].strip(), parts[1].strip()))
                    defs[sname] = struct_fields
    return defs


def main():
    parser = argparse.ArgumentParser(description='Migrate JSON configs to Excel format')
    parser.add_argument('--project-root', default=None)
    args = parser.parse_args()

    if args.project_root:
        project_root = args.project_root
    else:
        project_root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))

    json_dir = os.path.join(project_root, 'Assets', 'Resources', 'Configs')
    excel_dir = os.path.join(project_root, 'Configs')

    if not os.path.exists(excel_dir):
        os.makedirs(excel_dir)

    all_struct_defs = collect_struct_defs()

    # Load game_config.json for meta distribution
    game_config_path = os.path.join(json_dir, 'game_config.json')
    game_config = {}
    if os.path.exists(game_config_path):
        with codecs.open(game_config_path, 'r', encoding='utf-8-sig') as f:
            game_config = json.load(f)
        print("Loaded game_config.json: %s" % list(game_config.keys()))

    for config_name, schema in SCHEMAS.items():
        print("Migrating %s ..." % config_name)
        wb = Workbook()

        list_fields = schema.get('list', [])
        meta_fields = schema.get('meta', [])
        json_file = schema.get('json_file', config_name)
        json_root = schema.get('json_root', None)
        meta_from_gc = schema.get('meta_from_game_config', {})

        # Load JSON data
        json_path = os.path.join(json_dir, json_file + '.json')
        json_data = {}
        if os.path.exists(json_path):
            with codecs.open(json_path, 'r', encoding='utf-8-sig') as f:
                json_data = json.load(f)

        # --- List sheet ---
        if list_fields:
            ws = wb.active
            ws.title = config_name

            # Row 1: field names
            for ci, (fname, ftype) in enumerate(list_fields):
                ws.cell(row=1, column=ci + 1, value=fname)

            # Row 2: type declarations
            for ci, (fname, ftype) in enumerate(list_fields):
                ws.cell(row=2, column=ci + 1, value=ftype)

            # Row 3+: data
            items = []
            if json_root and json_root in json_data:
                items = json_data[json_root]
            elif isinstance(json_data, list):
                items = json_data

            for ri, item in enumerate(items):
                for ci, (fname, ftype) in enumerate(list_fields):
                    val = item.get(fname)
                    encoded = encode_value(val, ftype, all_struct_defs)
                    ws.cell(row=ri + 3, column=ci + 1, value=encoded)

            print("  List sheet: %d fields, %d rows" % (len(list_fields), len(items)))

        # --- Meta sheet ---
        if meta_fields:
            if list_fields:
                ws_meta = wb.create_sheet(title=config_name.replace('_config', '_meta').rstrip('_'))
            else:
                # Meta-only config (like global_config)
                ws_meta = wb.active
                short = config_name.replace('_config', '')
                ws_meta.title = short + '_meta' if short != config_name else 'meta'

            for ci, (fname, ftype) in enumerate(meta_fields):
                ws_meta.cell(row=1, column=ci + 1, value=fname)
            for ci, (fname, ftype) in enumerate(meta_fields):
                ws_meta.cell(row=2, column=ci + 1, value=ftype)

            # Get meta values from game_config
            for ci, (fname, ftype) in enumerate(meta_fields):
                val = None
                if fname in meta_from_gc:
                    gc_key = meta_from_gc[fname]
                    val = game_config.get(gc_key)
                if val is not None:
                    encoded = encode_value(val, ftype, all_struct_defs)
                    ws_meta.cell(row=3, column=ci + 1, value=encoded)

            print("  Meta sheet: %d fields" % len(meta_fields))

        # Remove default empty sheet if we created others
        if not list_fields and 'Sheet' in wb.sheetnames:
            del wb['Sheet']

        xlsx_path = os.path.join(excel_dir, config_name + '.xlsx')
        wb.save(xlsx_path)
        print("  Saved -> %s" % xlsx_path)

    print("\nMigration complete! %d Excel files created in %s" % (len(SCHEMAS), excel_dir))


if __name__ == '__main__':
    main()
